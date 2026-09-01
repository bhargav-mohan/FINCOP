from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import csv
import zipfile

from openpyxl import load_workbook

from finance_controller.ingestion.zipfile_extract import IngestError


def is_xlsx_package(path: str | Path) -> bool:
    src = Path(path)
    try:
        with zipfile.ZipFile(src) as zf:
            return any(name.replace("\\", "/").endswith("xl/workbook.xml") for name in zf.namelist())
    except zipfile.BadZipFile:
        return False


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def explode_xlsx(xlsx_path: str | Path, dest: str | Path) -> list[Path]:
    """Write one CSV per worksheet (sheet name as filename)."""
    src = Path(xlsx_path)
    out = Path(dest)
    out.mkdir(parents=True, exist_ok=True)
    try:
        wb = load_workbook(src, read_only=True, data_only=True)
    except Exception as exc:
        raise IngestError(f"could not read Excel workbook: {src.name}") from exc
    written: list[Path] = []
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_cell(h) or f"col_{i}" for i, h in enumerate(rows[0])]
        if not any(headers):
            continue
        target = out / f"{sheet_name}.csv"
        with target.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=headers)
            writer.writeheader()
            for raw in rows[1:]:
                row = {headers[i]: _cell(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
                if any(row.values()):
                    writer.writerow(row)
        written.append(target)
    if not written:
        raise IngestError(f"Excel workbook {src.name} has no data sheets")
    return written
